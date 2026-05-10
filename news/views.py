import csv
import io
import json
import os

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from .forms import NewsSubmissionForm
from .models import AnalysisResult, ModelMetrics, NewsArticle


def _require_verified(request):
    """Return a redirect response if the user is not verified, else None."""
    if not request.user.verified:
        messages.warning(
            request,
            'Your account is pending admin approval. '
            'Please wait until an administrator verifies your account.'
        )
        return redirect('login')
    return None


@login_required
def dashboard(request):
    guard = _require_verified(request)
    if guard:
        return guard

    all_articles = (
        NewsArticle.objects
        .filter(user=request.user)
        .select_related('result')
        .order_by('submitted_at')
    )
    recent_articles = list(all_articles)[-5:][::-1]
    latest_metrics = ModelMetrics.objects.first()
    model_ready = os.path.exists(
        os.path.join(settings.ML_MODEL_DIR, 'model.joblib')
    )

    # Chart data — Real/Fake distribution
    real_count = sum(1 for a in all_articles if getattr(a, 'result', None) and a.result.verdict == 'Real')
    fake_count = sum(1 for a in all_articles if getattr(a, 'result', None) and a.result.verdict == 'Fake')

    # Chart data — confidence trend (last 15 analyses)
    trend_articles = [a for a in all_articles if getattr(a, 'result', None)][-15:]
    trend_labels = [a.submitted_at.strftime('%b %d') for a in trend_articles]
    trend_confidence = [round(a.result.confidence_score, 1) for a in trend_articles]
    trend_verdicts = [a.result.verdict for a in trend_articles]

    return render(request, 'news/dashboard.html', {
        'recent_articles': recent_articles,
        'latest_metrics': latest_metrics,
        'model_ready': model_ready,
        'real_count': real_count,
        'fake_count': fake_count,
        'trend_labels': json.dumps(trend_labels),
        'trend_confidence': json.dumps(trend_confidence),
        'trend_verdicts': json.dumps(trend_verdicts),
    })


@login_required
def analyze(request):
    guard = _require_verified(request)
    if guard:
        return guard

    result_data = None
    form = NewsSubmissionForm()

    if request.method == 'POST':
        form = NewsSubmissionForm(request.POST)
        if form.is_valid():
            text = form.cleaned_data['text']
            model_dir = str(settings.ML_MODEL_DIR)

            try:
                from ml.predictor import predict
                prediction = predict(text, model_dir)
            except FileNotFoundError:
                messages.error(
                    request,
                    'The AI model has not been trained yet. '
                    'Please ask the administrator to upload a dataset and train the model.'
                )
                return render(request, 'news/analyze.html', {'form': form})
            except Exception as e:
                messages.error(request, f'Analysis failed: {e}')
                return render(request, 'news/analyze.html', {'form': form})

            # Persist to DB
            article = NewsArticle.objects.create(user=request.user, text=text)
            AnalysisResult.objects.create(
                article=article,
                verdict=prediction['verdict'],
                confidence_score=prediction['confidence'],
            )
            result_data = prediction

    return render(request, 'news/analyze.html', {
        'form': form,
        'result': result_data,
    })


@login_required
def history(request):
    guard = _require_verified(request)
    if guard:
        return guard

    articles = (
        NewsArticle.objects
        .filter(user=request.user)
        .select_related('result')
    )
    return render(request, 'news/history.html', {'articles': articles})


@login_required
def export_csv(request):
    guard = _require_verified(request)
    if guard:
        return guard

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="analysis_history.csv"'

    writer = csv.writer(response)
    writer.writerow(['#', 'Submitted At', 'Text (first 100 chars)', 'Verdict', 'Confidence (%)', 'Analyzed At'])

    articles = (
        NewsArticle.objects
        .filter(user=request.user)
        .select_related('result')
    )
    for i, article in enumerate(articles, 1):
        res = getattr(article, 'result', None)
        writer.writerow([
            i,
            article.submitted_at.strftime('%Y-%m-%d %H:%M'),
            article.text[:100].replace('\n', ' '),
            res.verdict if res else 'N/A',
            res.confidence_score if res else 'N/A',
            res.analyzed_at.strftime('%Y-%m-%d %H:%M') if res else 'N/A',
        ])

    return response


def _get_history_queryset(user):
    return (
        NewsArticle.objects
        .filter(user=user)
        .select_related('result')
    )


@login_required
def export_pdf(request):
    guard = _require_verified(request)
    if guard:
        return guard

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Analysis History', styles['Title']))
    elements.append(Paragraph(f'User: {request.user.get_full_name() or request.user.username}',
                               styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    headers = ['#', 'Submitted At', 'Article Preview', 'Verdict', 'Confidence (%)', 'Analyzed At']
    rows = [headers]

    for i, article in enumerate(_get_history_queryset(request.user), 1):
        res = getattr(article, 'result', None)
        rows.append([
            str(i),
            article.submitted_at.strftime('%Y-%m-%d %H:%M'),
            article.text[:60] + ('…' if len(article.text) > 60 else ''),
            res.verdict if res else 'N/A',
            f'{res.confidence_score:.2f}' if res else 'N/A',
            res.analyzed_at.strftime('%Y-%m-%d %H:%M') if res else 'N/A',
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="analysis_history.pdf"'
    return response


@login_required
def export_excel(request):
    guard = _require_verified(request)
    if guard:
        return guard

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Analysis History'

    headers = ['#', 'Submitted At', 'Article Preview', 'Verdict', 'Confidence (%)', 'Analyzed At']
    header_fill = PatternFill('solid', fgColor='1a1a2e')
    header_font = Font(bold=True, color='FFFFFF')

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, article in enumerate(_get_history_queryset(request.user), 1):
        res = getattr(article, 'result', None)
        ws.append([
            i,
            article.submitted_at.strftime('%Y-%m-%d %H:%M'),
            article.text[:100].replace('\n', ' '),
            res.verdict if res else 'N/A',
            round(res.confidence_score, 2) if res else 'N/A',
            res.analyzed_at.strftime('%Y-%m-%d %H:%M') if res else 'N/A',
        ])

    col_widths = [5, 18, 50, 10, 16, 18]
    for col, width in zip(ws.columns, col_widths):
        ws.column_dimensions[col[0].column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="analysis_history.xlsx"'
    return response


'''
register(request)

  Two scenarios depending on request type:
  - GET (user just visits the page) → create empty form → show registration page
  - POST (user submitted the form) → validate data → if valid, save user to DB → redirect to login with success message   

  form.is_valid() checks:
  - No empty required fields
  - Email is valid format
  - Username not already taken
  - Password1 matches password2

  profile(request)

  Same pattern:
  - GET → load form pre-filled with current user's data
  - POST → validate → save changes → redirect back to profile

  instance=request.user is key — it tells the form "this is an existing user, update them" instead of creating a new one. 

  @login_required
  — if someone tries to visit /accounts/profile/ without being logged in, Django automatically redirects them to the login   page. Without this decorator anyone could access the profile page.

  messages.success()
  — stores a one-time flash message that gets displayed on the next page (like "Registration successful!").
'''